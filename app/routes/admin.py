from functools import wraps
from io import BytesIO
from datetime import datetime, date

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app import mysql

admin_bp = Blueprint("admin", __name__)


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Access denied. Admin privileges required.", "danger")
            return redirect(url_for("main.home"))
        return f(*args, **kwargs)
    return decorated_function


@admin_bp.route("/dashboard")
@login_required
@admin_required
def dashboard():
    cur = mysql.connection.cursor()

    cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'user'")
    total_users = cur.fetchone()["count"]

    cur.execute("SELECT COUNT(*) as count FROM user_assessments WHERE status = 'completed'")
    total_assessments = cur.fetchone()["count"]

    cur.execute("SELECT COUNT(*) as count FROM career_recommendations")
    total_recommendations = cur.fetchone()["count"]

    cur.execute("SELECT COUNT(*) as count FROM contact_messages WHERE is_read = 0")
    unread_messages = cur.fetchone()["count"]

    cur.execute(
        "SELECT u.id, u.full_name, u.email, u.username, u.faculty, u.created_at "
        "FROM users u WHERE u.role = 'user' ORDER BY u.created_at DESC LIMIT 10"
    )
    recent_users = cur.fetchall()

    cur.execute(
        "SELECT cm.*, CASE WHEN cm.is_read = 0 THEN 'Unread' ELSE 'Read' END as status "
        "FROM contact_messages cm ORDER BY cm.created_at DESC LIMIT 5"
    )
    recent_messages = cur.fetchall()

    cur.close()

    return render_template(
        "admin/dashboard.html",
        total_users=total_users,
        total_assessments=total_assessments,
        total_recommendations=total_recommendations,
        unread_messages=unread_messages,
        recent_users=recent_users,
        recent_messages=recent_messages,
    )


@admin_bp.route("/users")
@login_required
@admin_required
def users():
    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT u.*, "
        "(SELECT COUNT(*) FROM user_assessments ua WHERE ua.user_id = u.id AND ua.status = 'completed') as assessments_count "
        "FROM users u WHERE u.role = 'user' ORDER BY u.created_at DESC"
    )
    all_users = cur.fetchall()
    cur.close()
    return render_template("admin/users.html", users=all_users)


@admin_bp.route("/users/<int:user_id>")
@login_required
@admin_required
def user_detail(user_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()

    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("admin.users"))

    cur.execute(
        "SELECT ua.*, "
        "(SELECT COUNT(*) FROM career_recommendations cr WHERE cr.assessment_id = ua.id) as recommendations_count "
        "FROM user_assessments ua WHERE ua.user_id = %s ORDER BY ua.started_at DESC",
        (user_id,),
    )
    assessments = cur.fetchall()

    cur.execute(
        "SELECT cr.*, c.title as career_title, cc.name as category_name "
        "FROM career_recommendations cr "
        "JOIN careers c ON cr.career_id = c.id "
        "JOIN career_categories cc ON c.category_id = cc.id "
        "WHERE cr.user_id = %s ORDER BY cr.match_score DESC",
        (user_id,),
    )
    recommendations = cur.fetchall()

    cur.close()
    return render_template(
        "admin/user_detail.html",
        user=user,
        assessments=assessments,
        recommendations=recommendations,
    )


@admin_bp.route("/messages")
@login_required
@admin_required
def messages():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM contact_messages ORDER BY created_at DESC")
    all_messages = cur.fetchall()
    cur.close()
    return render_template("admin/messages.html", messages=all_messages)


@admin_bp.route("/messages/<int:message_id>/read")
@login_required
@admin_required
def mark_message_read(message_id):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE contact_messages SET is_read = 1 WHERE id = %s", (message_id,))
    mysql.connection.commit()
    cur.close()
    flash("Message marked as read.", "success")
    return redirect(url_for("admin.messages"))


# ---------------------------------------------------------------------------
# Assessment Questions Management
# ---------------------------------------------------------------------------

@admin_bp.route("/questions")
@login_required
@admin_required
def questions():
    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT aq.*, "
        "(SELECT COUNT(*) FROM assessment_options ao WHERE ao.question_id = aq.id) AS options_count "
        "FROM assessment_questions aq ORDER BY aq.id"
    )
    all_questions = cur.fetchall()
    cur.close()
    return render_template("admin/questions.html", questions=all_questions)


@admin_bp.route("/questions/add", methods=["GET", "POST"])
@login_required
@admin_required
def add_question():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM career_categories ORDER BY name")
    categories = cur.fetchall()

    if request.method == "POST":
        question_text = request.form.get("question_text", "").strip()
        question_type = request.form.get("question_type")
        category = request.form.get("category", "").strip()
        weight = request.form.get("weight", 1, type=int)

        if not question_text or not question_type or not category:
            flash("Please fill in all required fields.", "danger")
            cur.close()
            return redirect(url_for("admin.add_question"))

        cur.execute(
            "INSERT INTO assessment_questions (question_text, question_type, category, weight) "
            "VALUES (%s, %s, %s, %s)",
            (question_text, question_type, category, weight),
        )
        question_id = cur.lastrowid

        option_texts = request.form.getlist("option_text[]")
        option_scores = request.form.getlist("option_score[]")
        option_categories = request.form.getlist("option_category[]")

        for i, text in enumerate(option_texts):
            text = text.strip()
            if not text:
                continue
            score = int(option_scores[i]) if i < len(option_scores) and option_scores[i] else 0
            cat_id = int(option_categories[i]) if i < len(option_categories) and option_categories[i] else None
            cur.execute(
                "INSERT INTO assessment_options (question_id, option_text, score, career_category_id) "
                "VALUES (%s, %s, %s, %s)",
                (question_id, text, score, cat_id),
            )

        mysql.connection.commit()
        cur.close()
        flash("Question added successfully!", "success")
        return redirect(url_for("admin.questions"))

    cur.close()
    return render_template("admin/question_form.html", question=None, categories=categories)


@admin_bp.route("/questions/<int:question_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit_question(question_id):
    cur = mysql.connection.cursor()

    cur.execute("SELECT * FROM assessment_questions WHERE id = %s", (question_id,))
    question = cur.fetchone()
    if not question:
        cur.close()
        flash("Question not found.", "danger")
        return redirect(url_for("admin.questions"))

    cur.execute("SELECT * FROM career_categories ORDER BY name")
    categories = cur.fetchall()

    if request.method == "POST":
        question_text = request.form.get("question_text", "").strip()
        question_type = request.form.get("question_type")
        category = request.form.get("category", "").strip()
        weight = request.form.get("weight", 1, type=int)

        if not question_text or not question_type or not category:
            flash("Please fill in all required fields.", "danger")
            cur.close()
            return redirect(url_for("admin.edit_question", question_id=question_id))

        cur.execute(
            "UPDATE assessment_questions SET question_text = %s, question_type = %s, "
            "category = %s, weight = %s WHERE id = %s",
            (question_text, question_type, category, weight, question_id),
        )

        cur.execute("DELETE FROM assessment_options WHERE question_id = %s", (question_id,))

        option_texts = request.form.getlist("option_text[]")
        option_scores = request.form.getlist("option_score[]")
        option_categories = request.form.getlist("option_category[]")

        for i, text in enumerate(option_texts):
            text = text.strip()
            if not text:
                continue
            score = int(option_scores[i]) if i < len(option_scores) and option_scores[i] else 0
            cat_id = int(option_categories[i]) if i < len(option_categories) and option_categories[i] else None
            cur.execute(
                "INSERT INTO assessment_options (question_id, option_text, score, career_category_id) "
                "VALUES (%s, %s, %s, %s)",
                (question_id, text, score, cat_id),
            )

        mysql.connection.commit()
        cur.close()
        flash("Question updated successfully!", "success")
        return redirect(url_for("admin.questions"))

    cur.execute("SELECT * FROM assessment_options WHERE question_id = %s ORDER BY id", (question_id,))
    options = cur.fetchall()
    cur.close()
    return render_template("admin/question_form.html", question=question, options=options, categories=categories)


@admin_bp.route("/questions/<int:question_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_question(question_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT id FROM assessment_questions WHERE id = %s", (question_id,))
    if not cur.fetchone():
        cur.close()
        flash("Question not found.", "danger")
        return redirect(url_for("admin.questions"))

    cur.execute("DELETE FROM assessment_questions WHERE id = %s", (question_id,))
    mysql.connection.commit()
    cur.close()
    flash("Question deleted successfully!", "success")
    return redirect(url_for("admin.questions"))


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

def _parse_date_filters():
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    parsed_from = None
    parsed_to = None
    if date_from:
        try:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass
    if date_to:
        try:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass
    return date_from, date_to, parsed_from, parsed_to


def _fetch_report_data(report_type, parsed_from, parsed_to):
    """Return (headers, rows, title) for a given report type and date range."""
    cur = mysql.connection.cursor()

    if report_type == "students":
        sql = (
            "SELECT u.id, u.full_name, u.email, u.username, u.gender, "
            "u.faculty, u.department, u.year_of_study, u.created_at, "
            "(SELECT COUNT(*) FROM user_assessments ua WHERE ua.user_id = u.id AND ua.status = 'completed') AS assessments "
            "FROM users u WHERE u.role = 'user'"
        )
        params = []
        if parsed_from:
            sql += " AND DATE(u.created_at) >= %s"
            params.append(parsed_from)
        if parsed_to:
            sql += " AND DATE(u.created_at) <= %s"
            params.append(parsed_to)
        sql += " ORDER BY u.created_at DESC"
        cur.execute(sql, params)
        rows_raw = cur.fetchall()
        headers = ["ID", "Full Name", "Email", "Username", "Gender", "Faculty", "Department", "Year", "Registered", "Assessments"]
        rows = [
            [r["id"], r["full_name"], r["email"], r["username"], r["gender"] or "N/A",
             r["faculty"] or "N/A", r["department"] or "N/A", r["year_of_study"] or "N/A",
             r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "N/A", r["assessments"]]
            for r in rows_raw
        ]
        title = "Students Report"

    elif report_type == "assessments":
        sql = (
            "SELECT ua.id, u.full_name, u.email, ua.status, ua.started_at, ua.completed_at, "
            "(SELECT COUNT(*) FROM career_recommendations cr WHERE cr.assessment_id = ua.id) AS recommendations "
            "FROM user_assessments ua "
            "JOIN users u ON ua.user_id = u.id"
        )
        params = []
        conditions = []
        if parsed_from:
            conditions.append("DATE(ua.started_at) >= %s")
            params.append(parsed_from)
        if parsed_to:
            conditions.append("DATE(ua.started_at) <= %s")
            params.append(parsed_to)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY ua.started_at DESC"
        cur.execute(sql, params)
        rows_raw = cur.fetchall()
        headers = ["ID", "Student", "Email", "Status", "Started", "Completed", "Recommendations"]
        rows = [
            [r["id"], r["full_name"], r["email"], r["status"],
             r["started_at"].strftime("%Y-%m-%d %H:%M") if r["started_at"] else "N/A",
             r["completed_at"].strftime("%Y-%m-%d %H:%M") if r["completed_at"] else "N/A",
             r["recommendations"]]
            for r in rows_raw
        ]
        title = "Assessments Report"

    elif report_type == "recommendations":
        sql = (
            "SELECT cr.id, u.full_name, u.email, c.title AS career, cc.name AS category, "
            "cr.match_score, cr.created_at "
            "FROM career_recommendations cr "
            "JOIN users u ON cr.user_id = u.id "
            "JOIN careers c ON cr.career_id = c.id "
            "JOIN career_categories cc ON c.category_id = cc.id"
        )
        params = []
        conditions = []
        if parsed_from:
            conditions.append("DATE(cr.created_at) >= %s")
            params.append(parsed_from)
        if parsed_to:
            conditions.append("DATE(cr.created_at) <= %s")
            params.append(parsed_to)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY cr.created_at DESC"
        cur.execute(sql, params)
        rows_raw = cur.fetchall()
        headers = ["ID", "Student", "Email", "Career", "Category", "Match Score (%)", "Date"]
        rows = [
            [r["id"], r["full_name"], r["email"], r["career"], r["category"],
             r["match_score"],
             r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "N/A"]
            for r in rows_raw
        ]
        title = "Career Recommendations Report"

    else:  # messages
        sql = "SELECT id, name, email, subject, message, is_read, created_at FROM contact_messages"
        params = []
        conditions = []
        if parsed_from:
            conditions.append("DATE(created_at) >= %s")
            params.append(parsed_from)
        if parsed_to:
            conditions.append("DATE(created_at) <= %s")
            params.append(parsed_to)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY created_at DESC"
        cur.execute(sql, params)
        rows_raw = cur.fetchall()
        headers = ["ID", "Name", "Email", "Subject", "Message", "Status", "Date"]
        rows = [
            [r["id"], r["name"], r["email"], r["subject"],
             (r["message"][:80] + "...") if len(r["message"]) > 80 else r["message"],
             "Read" if r["is_read"] else "Unread",
             r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "N/A"]
            for r in rows_raw
        ]
        title = "Contact Messages Report"

    cur.close()
    return headers, rows, title


@admin_bp.route("/reports")
@login_required
@admin_required
def reports():
    report_type = request.args.get("type", "students")
    date_from, date_to, parsed_from, parsed_to = _parse_date_filters()
    headers, rows, title = _fetch_report_data(report_type, parsed_from, parsed_to)

    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM users WHERE role = 'user'")
    total_students = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) AS c FROM user_assessments WHERE status = 'completed'")
    total_assessments = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) AS c FROM career_recommendations")
    total_recommendations = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) AS c FROM contact_messages")
    total_messages = cur.fetchone()["c"]
    cur.close()

    return render_template(
        "admin/reports.html",
        report_type=report_type,
        date_from=date_from,
        date_to=date_to,
        headers=headers,
        rows=rows,
        title=title,
        total_students=total_students,
        total_assessments=total_assessments,
        total_recommendations=total_recommendations,
        total_messages=total_messages,
    )


@admin_bp.route("/reports/download/excel")
@login_required
@admin_required
def download_excel():
    report_type = request.args.get("type", "students")
    _, _, parsed_from, parsed_to = _parse_date_filters()
    headers, rows, title = _fetch_report_data(report_type, parsed_from, parsed_to)

    wb = Workbook()
    ws = wb.active
    ws.title = title

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"MKU AI Career Guidance — {title}")
    title_cell.font = Font(bold=True, size=14, color="1F2937")
    title_cell.alignment = Alignment(horizontal="center")

    date_label = "Generated: " + datetime.now().strftime("%B %d, %Y at %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=date_label).font = Font(italic=True, color="6B7280")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (row_idx - 5) % 2 == 1:
                cell.fill = alt_fill

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_data in rows:
            val_len = len(str(row_data[col_idx - 1])) if col_idx - 1 < len(row_data) else 0
            max_length = max(max_length, val_len)
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_length + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_type}_report_{date.today().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/reports/download/pdf")
@login_required
@admin_required
def download_pdf():
    report_type = request.args.get("type", "students")
    _, _, parsed_from, parsed_to = _parse_date_filters()
    headers, rows, title = _fetch_report_data(report_type, parsed_from, parsed_to)

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"],
                                 fontSize=18, textColor=colors.HexColor("#1F2937"),
                                 spaceAfter=4)
    subtitle_style = ParagraphStyle("ReportSub", parent=styles["Normal"],
                                    fontSize=10, textColor=colors.HexColor("#6B7280"),
                                    spaceAfter=16)

    elements = []
    elements.append(Paragraph(f"MKU AI Career Guidance — {title}", title_style))
    elements.append(Paragraph("Generated: " + datetime.now().strftime("%B %d, %Y at %H:%M"), subtitle_style))
    elements.append(Spacer(1, 8))

    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8,
                                leading=10, wordWrap="CJK")

    table_data = [headers]
    for row in rows:
        table_data.append([Paragraph(str(v), cell_style) for v in row])

    col_count = len(headers)
    available_width = landscape(A4)[0] - 1 * inch
    col_width = available_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 16))
    footer_style = ParagraphStyle("Footer", parent=styles["Normal"],
                                  fontSize=8, textColor=colors.HexColor("#9CA3AF"),
                                  alignment=1)
    elements.append(Paragraph(f"Total records: {len(rows)} | © {datetime.now().year} Mount Kigali University", footer_style))

    doc.build(elements)
    output.seek(0)

    filename = f"{report_type}_report_{date.today().isoformat()}.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")
